import openpyxl
import easygui
import pickle
from openpyxl.comments import Comment
from openpyxl import styles
import random
from os.path import basename
from myapp.Schedule.remark import Remark


def main(excel):
    excel.create_sheet("统计总表")  # 创建一个工作表储存日志
    num = 1
    for sheet_name in excel.sheetnames:  # 遍历课表
        for n1 in range(2, 7):
            for n2 in range(4, 12):
                row = convert2title(n1)  # 数字转换为对应的列
                cell = row + str(n2)  # 单元格具体坐标
                if excel[sheet_name][cell].value is not None:
                    teacher_name = excel[sheet_name][cell].value.replace('\r', '').split("\n")[1]  # 单元格中教师名
                    if teacher_name in person_name:  # 如果教师名字存在于请假列表中
                        info = teacher_info.copy()  # 教师信息备份

                        subject_name = excel[sheet_name][cell].value.replace('\r', '').split("\n")[
                            0]  # 单元格中科目名
                        grade = sheet_name[:3]  # 年级
                        sheet_class = sheet_name[4]  # 班级
                        if way == "1、课表自动修改(优先科目)":
                            info1 = filter_1(subject_name, info)  # 能上指定课程的教师信息字典
                            info2 = filter_2(cell, excel, info1)  # 能上指定课程并且课程不冲突的教师信息
                            if len(info2) != 0:
                                final_info = init_grade(grade, info2)  # 最终的教师名单
                                final_name = random_name(final_info)  # 代课老师
                                comment(excel[sheet_name][cell], teacher_name, final_name)
                                excel[sheet_name][cell].value = subject_name + '\n' + final_name
                                cell_format(excel[sheet_name][cell])
                                log(num, excel, teacher_name, subject_name, sheet_name, cell, final_info[0])
                                num += 1
                            else:
                                excel[sheet_name][cell].value = subject_name + '\n' + "无候选教师"
                                cell_format1(excel[sheet_name][cell])
                                log(num, excel, teacher_name, subject_name, sheet_name, row + str(n2))
                                referrals(teacher_info, grade, excel, cell, num)
                                num += 1
                        elif way == "2、课表自动修改(优先班级)":
                            info1 = filter_grade_class(info, grade, sheet_class)
                            info2 = filter_2(cell, excel, info1)
                            if len(info2) != 0:
                                final_name = random_name(info2)
                                comment(excel[sheet_name][cell], teacher_name, final_name)
                                cache = random_name(teacher_info[final_name][1].split('、'))
                                excel[sheet_name][cell].value = cache + '\n' + final_name
                                cell_format(excel[sheet_name][cell])
                                log(num, excel, teacher_name, subject_name, sheet_name, row + str(n2), final_name)
                                num += 1
                            else:
                                excel[sheet_name][cell].value = subject_name + '\n' + "无候选教师"
                                cell_format1(excel[sheet_name][cell])
                                log(num, excel, teacher_name, subject_name, sheet_name, row + str(n2))
                                referrals(teacher_info, grade, excel, cell, num)
                                num += 1
    excel.save(file_save)


def filter_grade_class(teachers_info, grade, sheet_class):
    """筛选年级、班级符合的教师"""
    cache = []
    for each in teachers_info.keys():
        if teachers_info[each][0] == grade and sheet_class in teachers_info[each][2]:
            cache.append(each)
    return cache


def referrals(info, grade, excel, cell, num):
    """无代课老师时添加推荐人员"""
    cache = []
    for each in info:
        if grade in info[each][0]:
            cache.append(each)
    for sheet_name in excel.sheetnames:
        if excel[sheet_name][cell].value is not None:
            if excel[sheet_name][cell].value.replace('\r', '').split("\n")[1] in cache:
                cache.remove(excel[sheet_name][cell].value.replace('\r', '').split("\n")[1])
    if len(cache) != 0:
        output = ""
        for n in cache:
            output += f"班级：{teacher_info[n][0]}  {teacher_info[n][2]}   姓名：{n} 科目：{teacher_info[n][1]}\n"
            comm = Comment(output, "")
            comm.width, comm.height = 600, 500
            excel["统计总表"]["A" + str(num)].comment = comm


def log(num, excel, old_name, subject, sheet, cell, new_name="无"):
    """日志"""
    excel["统计总表"].merge_cells(f"A{num}:H{num}")
    excel["统计总表"]["A" + str(num)].hyperlink = f"{basename(file_save)}#'{sheet}'!{cell}"  # 超链接
    excel["统计总表"][
        "A" + str(num)].value = f"班级：{sheet}    科目：{subject}    任课老师：{old_name}    代课老师：{new_name} "


def random_name(teacher_name):
    """在候选名单中随机选一名教师代课"""
    end = len(teacher_name) - 1
    return teacher_name[random.randint(0, end)]


def comment(cell, old_teacher, new_teacher):
    """修改过的表格自动填入批注"""
    cache = Comment(f"初始教师：{old_teacher}\n代课教师：{new_teacher}", "")
    cache.width, cache.height = 120, 40
    cell.comment = cache


def cell_format(cell):
    """修改单元格字体颜色"""
    cell.font = styles.Font(color='0099CC00')  # 有代课教师填充绿色


def cell_format1(cell):
    """修改单元格字体颜色"""
    cell.font = styles.Font(color='003366FF')  # 无代课教师填充蓝色


def init_grade(object_grade, info):
    """筛选出最终候选名单"""
    cache = []
    for each in info:
        if object_grade in info[each][0]:
            cache.append(each)
    if len(cache) == 0:  # 没有找到对应年级的老师
        for each in info:
            if object_grade == "一年级":
                if "二年级" in info[each][0]:
                    cache.append(each)
            elif object_grade == "二年级":
                if "一年级" in info[each][0]:
                    cache.append(each)
            elif object_grade == "七年级":
                if "八年级" in info[each][0]:
                    cache.append(each)
            elif object_grade == "八年级":
                if "七年级" in info[each][0]:
                    cache.append(each)
    if len(cache) == 0:  # 没有相近年级的教师
        for each in info:
            cache.append(each)

    return cache


def filter_1(subject, info):
    """删除教师信息中科目不相关人员"""
    if subject == '校本-英语':
        subject = '英语'
    cache = list()
    for each in info:
        check = 0
        for each2 in info[each][1].split('、'):
            if each2 == subject:
                check = 1
        if check == 0:
            cache.append(each)
    for each in cache:
        info.pop(each)
    return info


def filter_2(cell, excel, info):
    """便利每个单元格的对应单元格筛选出闲置教师"""
    for sheet_name in excel.sheetnames:
        if excel[sheet_name][cell].value is not None:
            if excel[sheet_name][cell].value.replace('\r', '').split("\n")[1] in info:
                info.pop(info.index(excel[sheet_name][cell].value.replace('\r', '').split("\n")[1]))
    return info


def convert2title(n):
    """将数字转换成EXCEL对应的字母列"""
    title = ""
    while n != 0:
        res = n % 26
        if res == 0:
            res = 26
            n -= 26
        title = chr(ord('A') + res - 1) + title
        n = n // 26
    return title


def load_pickle(file):
    """加载pickle文件"""
    with open(file=file, mode='rb') as cache:
        return pickle.load(cache)


def remove_person(info, name):
    """删除字典中存在的person信息"""
    for each in name:
        try:
            info.pop(each)
        except KeyError:
            print(f"{each}不存在老师信息中")
            choose = input("是否继续执行程序？ y/n")
            if choose.lower() == "n":
                raise KeyError
    return info


if __name__ == '__main__':
    way = easygui.choicebox("选择功能", "选择功能", ["1、课表自动修改(优先科目)", "2、课表自动修改(优先班级)", "3、根据总表改内容"])
    if way == "1、课表自动修改(优先科目)" or "2、课表自动修改(优先班级)":
        teacher_info = load_pickle("teacher_info.pickle")

        with open("请假人员.txt", mode='r') as temp:
            person_name = temp.readline().split('、')

        teacher_info = remove_person(teacher_info, person_name)

        easygui.msgbox("请选择班级课表所在位置")
        teacher_schedule = easygui.fileopenbox("选择文件路径", default=r"C:\Users\Administrator\Desktop\*.xlsx",
                                               filetypes=["*.xlsx"])
        easygui.msgbox("请选择汇总文件保存位置")
        file_save = easygui.filesavebox(default=r"C:\Users\Administrator\Desktop\*.xlsx",
                                        filetypes=["*.xlsx"])

        teacher_schedule_work = openpyxl.load_workbook(teacher_schedule)
        main(teacher_schedule_work)
    elif way == "3、根据总表改内容":
        r = Remark()
